import os
import json
import logging
import xml.etree.ElementTree as ET
from xml.dom import minidom
from typing import Dict, Any
import pandas as pd
from src.agents.base import BaseAgent
from src.state import GraphState

logger = logging.getLogger(__name__)

class OutputGeneratorAgent(BaseAgent):
    """
    Agent responsible for generating client deliverables (Excel, CSV, JSON, XML)
    and exporting translations to SPPID, AVEVA, COMOS, and SP3D.
    """
    def run(self, state: GraphState) -> Dict[str, Any]:
        logger.info("Running Output Generator Agent...")
        
        graph = state.get("engineering_graph")
        if not graph:
            raise ValueError("No compiled UniversalEngineeringGraph found in state to generate deliverables.")
            
        output_dir = os.path.join(os.getcwd(), "outputs")
        os.makedirs(output_dir, exist_ok=True)
        
        deliverables_paths = {}
        
        # 1. Generate Excel Deliverable
        excel_path = os.path.join(output_dir, "engineering_deliverables.xlsx")
        self._generate_excel(graph, excel_path)
        deliverables_paths["excel"] = excel_path
        
        # 2. Generate Master JSON Graph Deliverable
        json_path = os.path.join(output_dir, "master_graph.json")
        with open(json_path, "w") as f:
            json.dump(graph.model_dump(), f, indent=2)
        deliverables_paths["json_graph"] = json_path
        
        # 3. Generate XML export for AVEVA Diagrams & SP3D
        aveva_path = os.path.join(output_dir, "aveva_diagrams_export.xml")
        self._generate_aveva_xml(graph, aveva_path)
        deliverables_paths["aveva_xml"] = aveva_path
        
        # 4. Generate COMOS Hierarchical JSON
        comos_path = os.path.join(output_dir, "comos_hierarchy_export.json")
        self._generate_comos_json(graph, comos_path)
        deliverables_paths["comos_json"] = comos_path
        
        # 5. Generate SPPID database table CSVs
        sppid_path = os.path.join(output_dir, "sppid_import_tables.csv")
        self._generate_sppid_csv(graph, sppid_path)
        deliverables_paths["sppid_csv"] = sppid_path

        logger.info(f"Successfully generated all deliverables inside: '{output_dir}'")
        
        # Cleanup uploaded file from Google GenAI File API to free up resources
        file_name = state.get("metadata", {}).get("primary_page_name")
        if file_name:
            try:
                from google import genai
                from src.config import GEMINI_API_KEY
                if GEMINI_API_KEY:
                    logger.info(f"Cleaning up uploaded file '{file_name}' from Google GenAI Files API...")
                    client = genai.Client(api_key=GEMINI_API_KEY)
                    client.files.delete(name=file_name)
                    logger.info("Cloud file deletion completed successfully.")
            except Exception as e:
                logger.warning(f"Could not delete uploaded file '{file_name}' from Google GenAI Files API: {e}")
        
        return {
            "deliverables": deliverables_paths,
            "revision_history": state.get("revision_history", []) + [{"action": "Generated final engineering deliverables and platform export maps"}]
        }

    def _generate_excel(self, graph, path: str):
        """
        Creates a styled Excel workbook with sheets for each detected entity type.
        Adapts dynamically to drawing type — works for P&ID, Electrical, Earthing, SLD, etc.
        """
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = openpyxl.Workbook()
        default_sheet = wb.active
        wb.remove(default_sheet)

        # Style definitions
        title_font = Font(name="Segoe UI", size=13, bold=True, color="FFFFFF")
        sub_font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
        header_font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
        cell_font = Font(name="Segoe UI", size=10)
        bold_cell_font = Font(name="Segoe UI", size=10, bold=True)

        banner_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
        sub_banner_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        alt_fill = PatternFill(start_color="EEF3FA", end_color="EEF3FA", fill_type="solid")

        align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9'),
        )

        drawing_type = getattr(graph, 'drawing_type', 'GENERIC').upper()
        dt_label = drawing_type.replace('_', ' ')

        def write_sheet(sheet_name: str, title: str, subtitle: str, headers: list, rows: list):
            ws = wb.create_sheet(title=sheet_name[:31])  # Excel max 31 chars
            # Title banner
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(headers), 1))
            c = ws.cell(row=1, column=1, value=title)
            c.font = title_font; c.fill = banner_fill; c.alignment = align_center
            ws.row_dimensions[1].height = 32
            for col in range(2, len(headers) + 1):
                ws.cell(row=1, column=col).fill = banner_fill
            # Subtitle
            ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max(len(headers), 1))
            c2 = ws.cell(row=2, column=1, value=subtitle)
            c2.font = sub_font; c2.fill = sub_banner_fill; c2.alignment = align_center
            ws.row_dimensions[2].height = 22
            for col in range(2, len(headers) + 1):
                ws.cell(row=2, column=col).fill = sub_banner_fill
            # Headers
            for ci, h in enumerate(headers, 1):
                c = ws.cell(row=3, column=ci, value=h)
                c.font = header_font; c.fill = header_fill
                c.alignment = align_center; c.border = thin_border
            ws.row_dimensions[3].height = 24
            # Data rows
            for ri, row_dict in enumerate(rows, 4):
                ws.row_dimensions[ri].height = 18
                fill = alt_fill if ri % 2 == 0 else None
                for ci, h in enumerate(headers, 1):
                    val = row_dict.get(h, "")
                    c = ws.cell(row=ri, column=ci, value=val)
                    c.font = bold_cell_font if ci == 1 else cell_font
                    c.border = thin_border
                    c.alignment = align_center if ci <= 2 else align_left
                    if fill:
                        c.fill = fill
            # Auto-fit
            for col in ws.columns:
                max_len = max((len(str(c.value or '')) for c in col), default=10)
                ws.column_dimensions[openpyxl.utils.get_column_letter(col[0].column)].width = min(max_len + 4, 50)

        # ── Build sheets based on drawing type ───────────────────────────────
        if drawing_type in ('PID', 'PFD', 'ISOMETRIC'):
            # Line List
            if graph.lines:
                headers = ["#", "Line Tag", "Size", "Service Code", "Sequence No.", "Pipe Spec", "Insulation", "From", "To"]
                rows = [{
                    "#": i + 1,
                    "Line Tag": l.tag, "Size": l.size, "Service Code": l.service,
                    "Sequence No.": l.sequence_number, "Pipe Spec": l.spec,
                    "Insulation": l.insulation or "-",
                    "From": l.from_node or "-", "To": l.to_node or "-",
                } for i, l in enumerate(graph.lines)]
                write_sheet("Line List", f"LINE LIST — {dt_label}", f"Extracted by SID-AI | {len(rows)} lines detected", headers, rows)

            # Instrument List
            if graph.instruments:
                headers = ["#", "Instrument Tag", "Type", "Loop ID", "Location", "Service"]
                rows = [{
                    "#": i + 1,
                    "Instrument Tag": inst.tag, "Type": inst.type,
                    "Loop ID": inst.loop_id or "-", "Location": inst.location or "Field",
                    "Service": inst.service or "-",
                } for i, inst in enumerate(graph.instruments)]
                write_sheet("Instrument List", f"INSTRUMENT LIST — {dt_label}", f"Extracted by SID-AI | {len(rows)} instruments detected", headers, rows)

            # Valve List
            if graph.valves:
                headers = ["#", "Valve Tag", "Type", "Size", "Rating", "Normal State", "Associated Line"]
                rows = [{
                    "#": i + 1,
                    "Valve Tag": v.tag, "Type": v.type, "Size": v.size or "-",
                    "Rating": v.rating or "-", "Normal State": v.normal_state or "-",
                    "Associated Line": v.line_tag or "-",
                } for i, v in enumerate(graph.valves)]
                write_sheet("Valve List", f"MANUAL VALVE LIST — {dt_label}", f"Extracted by SID-AI | {len(rows)} valves detected", headers, rows)

            # PSV List
            if graph.safety_relief_valves:
                headers = ["#", "PSV Tag", "Type", "Service", "Unit", "Set Pressure", "Inlet Size", "Outlet Size", "Inlet Spec", "Relief Destination", "Remarks"]
                rows = [{
                    "#": i + 1,
                    "PSV Tag": p.tag, "Type": p.type, "Service": p.service, "Unit": p.unit,
                    "Set Pressure": p.set_pressure, "Inlet Size": p.inlet_size,
                    "Outlet Size": p.outlet_size, "Inlet Spec": p.inlet_spec,
                    "Relief Destination": p.relief_destination, "Remarks": p.remarks or "-",
                } for i, p in enumerate(graph.safety_relief_valves)]
                write_sheet("PSV List", f"SAFETY RELIEF VALVE LIST — {dt_label}", f"Extracted by SID-AI | {len(rows)} PSVs detected", headers, rows)

            # Equipment List
            if graph.equipment:
                headers = ["#", "Equipment Tag", "Type", "Description", "Design Pressure", "Design Temp", "Flow Rate", "Duty", "Material"]
                rows = [{
                    "#": i + 1,
                    "Equipment Tag": e.tag, "Type": e.type, "Description": e.description or e.name,
                    "Design Pressure": e.design_pressure or "-", "Design Temp": e.design_temperature or "-",
                    "Flow Rate": e.flow_rate or "-", "Duty": e.duty or "-", "Material": e.material or "-",
                } for i, e in enumerate(graph.equipment)]
                write_sheet("Equipment List", f"EQUIPMENT LIST — {dt_label}", f"Extracted by SID-AI | {len(rows)} equipment items detected", headers, rows)

        elif drawing_type == 'ELECTRICAL_LAYOUT':
            if graph.luminaires:
                headers = ["#", "Luminaire Tag", "Fitting Type", "Wattage", "Circuit", "Panel", "Elevation", "Location"]
                rows = [{
                    "#": i + 1, "Luminaire Tag": l.tag, "Fitting Type": l.fitting_type or "-",
                    "Wattage": l.wattage or "-", "Circuit": l.circuit or "-", "Panel": l.panel or "-",
                    "Elevation": l.elevation or "-", "Location": l.location or "-",
                } for i, l in enumerate(graph.luminaires)]
                write_sheet("Luminaires", "LUMINAIRE / FITTING LIST", f"Extracted by SID-AI | {len(rows)} luminaires", headers, rows)

            if graph.panels:
                headers = ["#", "Panel Tag", "Type", "Voltage", "Capacity (kVA)", "Fed From", "Location", "Description"]
                rows = [{
                    "#": i + 1, "Panel Tag": p.tag, "Type": p.panel_type or "-",
                    "Voltage": p.voltage or "-", "Capacity (kVA)": p.capacity_kva or "-",
                    "Fed From": p.feeder_from or "-", "Location": p.location or "-",
                    "Description": p.description or "-",
                } for i, p in enumerate(graph.panels)]
                write_sheet("Panels / DBs", "DISTRIBUTION BOARDS & PANELS", f"Extracted by SID-AI | {len(rows)} panels", headers, rows)

            if graph.cables:
                headers = ["#", "Circuit Tag", "Cable Type", "Size (mm²)", "Cores", "From Panel", "To Equipment", "Route"]
                rows = [{
                    "#": i + 1, "Circuit Tag": c.tag, "Cable Type": c.cable_type or "-",
                    "Size (mm²)": c.size_mm2 or "-", "Cores": c.cores or "-",
                    "From Panel": c.from_panel or "-", "To Equipment": c.to_equipment or "-",
                    "Route": c.route or "-",
                } for i, c in enumerate(graph.cables)]
                write_sheet("Cables", "CABLE & CIRCUIT SCHEDULE", f"Extracted by SID-AI | {len(rows)} cables", headers, rows)

        elif drawing_type == 'EARTHING_LAYOUT':
            if graph.earthing_components:
                headers = ["#", "Tag", "Component Type", "Material", "Size", "Connected To", "Location", "Resistance"]
                rows = [{
                    "#": i + 1, "Tag": e.tag, "Component Type": e.component_type,
                    "Material": e.material or "-", "Size": e.size or "-",
                    "Connected To": e.connected_to or "-", "Location": e.location or "-",
                    "Resistance": e.resistance or "-",
                } for i, e in enumerate(graph.earthing_components)]
                write_sheet("Earthing", "EARTHING COMPONENT SCHEDULE", f"Extracted by SID-AI | {len(rows)} earthing items", headers, rows)

            if graph.equipment:
                headers = ["#", "Equipment Tag", "Type", "Description", "Location"]
                rows = [{
                    "#": i + 1, "Equipment Tag": e.tag, "Type": e.type,
                    "Description": e.description or e.name or "Earthed Footprint",
                    "Location": getattr(e, 'location', None) or "Grid Footprint",
                } for i, e in enumerate(graph.equipment)]
                write_sheet("Earthed Equipment", "EARTHED EQUIPMENT & STRUCTURAL FOOTPRINTS", f"Extracted by SID-AI | {len(rows)} earthed items", headers, rows)

            if graph.relationships:
                headers = ["#", "Source (From)", "Relation Type", "Target (Connected To)", "Confidence"]
                rows = [{
                    "#": i + 1,
                    "Source (From)": getattr(r, 'source_tag', getattr(r, 'source', '')),
                    "Relation Type": str(getattr(r, 'rel_type', getattr(r, 'type', ''))).upper(),
                    "Target (Connected To)": getattr(r, 'target_tag', getattr(r, 'target', '')),
                    "Confidence": f"{int(getattr(r, 'confidence', 1.0) * 100)}%",
                } for i, r in enumerate(graph.relationships)]
                write_sheet("Earthing Connectivity", "EARTHING TOPOLOGICAL RELATIONS (EARTHED_TO)", f"Extracted by SID-AI | {len(rows)} relations", headers, rows)

        elif drawing_type == 'SLD':
            if graph.panels:
                headers = ["#", "Panel Tag", "Type", "Voltage", "Capacity (kVA)", "Fed From", "Description"]
                rows = [{
                    "#": i + 1, "Panel Tag": p.tag, "Type": p.panel_type or "-",
                    "Voltage": p.voltage or "-", "Capacity (kVA)": p.capacity_kva or "-",
                    "Fed From": p.feeder_from or "-", "Description": p.description or "-",
                } for i, p in enumerate(graph.panels)]
                write_sheet("Switchgear", "SWITCHGEAR & PANEL SCHEDULE", f"Extracted by SID-AI | {len(rows)} panels", headers, rows)

            if graph.cables:
                headers = ["#", "Feeder Tag", "Cable Type", "Size (mm²)", "Cores", "From Panel", "To Equipment", "Route"]
                rows = [{
                    "#": i + 1, "Feeder Tag": c.tag, "Cable Type": c.cable_type or "-",
                    "Size (mm²)": c.size_mm2 or "-", "Cores": c.cores or "-",
                    "From Panel": c.from_panel or "-", "To Equipment": c.to_equipment or "-",
                    "Route": c.route or "-",
                } for i, c in enumerate(graph.cables)]
                write_sheet("Feeders", "FEEDER & BREAKER SCHEDULE", f"Extracted by SID-AI | {len(rows)} feeders", headers, rows)

            if graph.relationships:
                headers = ["#", "Source (From)", "Relation Type", "Target (Connected To)", "Confidence"]
                rows = [{
                    "#": i + 1,
                    "Source (From)": getattr(r, 'source_tag', getattr(r, 'source', '')),
                    "Relation Type": str(getattr(r, 'rel_type', getattr(r, 'type', ''))).upper(),
                    "Target (Connected To)": getattr(r, 'target_tag', getattr(r, 'target', '')),
                    "Confidence": f"{int(getattr(r, 'confidence', 1.0) * 100)}%",
                } for i, r in enumerate(graph.relationships)]
                write_sheet("Electrical Relations", "SLD POWER FLOW RELATIONS (FEEDS)", f"Extracted by SID-AI | {len(rows)} relations", headers, rows)

        else:
            # Generic fallback
            all_items = (
                [("EQUIPMENT", e.tag, e.type, e.description or "") for e in graph.equipment] +
                [("COMPONENT", g.tag, g.classification, g.description or "") for g in graph.generic_components]
            )
            if all_items:
                headers = ["#", "Category", "Tag", "Type", "Description"]
                rows = [{"#": i + 1, "Category": a[0], "Tag": a[1], "Type": a[2], "Description": a[3]} for i, a in enumerate(all_items)]
                write_sheet("Components", f"COMPONENT LIST — {dt_label}", f"Extracted by SID-AI | {len(rows)} items", headers, rows)

        # Always add Annotations sheet if present
        if graph.annotations:
            headers = ["#", "Type", "Text", "Position X", "Position Y"]
            rows = [{
                "#": i + 1, "Type": a.annotation_type, "Text": a.text,
                "Position X": round(a.position_x, 3) if a.position_x is not None else "-",
                "Position Y": round(a.position_y, 3) if a.position_y is not None else "-",
            } for i, a in enumerate(graph.annotations)]
            write_sheet("Annotations", "ANNOTATIONS & NOTES", f"Extracted by SID-AI | {len(rows)} annotations", headers, rows)

        if not wb.sheetnames:
            ws = wb.create_sheet("No Data")
            ws.cell(row=1, column=1, value="No entities were detected from this drawing.")

        wb.save(path)


    def _generate_aveva_xml(self, graph, path: str):
        """
        Exports the engineering objects to an AVEVA Diagrams / SP3D compatible XML document.
        Ensures all XML attributes and text nodes are valid non-null strings for Python 3.13 compatibility.
        """
        root = ET.Element("PlantModel", SchemaVersion="1.0", Source="SID-AI")
        
        # 1. Equipment Node
        if graph.equipment:
            eq_group = ET.SubElement(root, "EquipmentGroup")
            for eq in graph.equipment:
                el = ET.SubElement(eq_group, "Equipment", Tag=str(eq.tag or ""), Type=str(eq.type or ""))
                ET.SubElement(el, "Description").text = str(eq.description or "")
                ET.SubElement(el, "DesignPressure").text = str(eq.design_pressure or "")
                ET.SubElement(el, "DesignTemperature").text = str(eq.design_temperature or "")
                ET.SubElement(el, "Duty").text = str(eq.duty or "")
            
        # 2. Piping Group Node
        if graph.lines:
            pipe_group = ET.SubElement(root, "PipingNetwork")
            for line in graph.lines:
                el = ET.SubElement(pipe_group, "Pipeline", Tag=str(line.tag or ""), Size=str(line.size or ""), Spec=str(line.spec or ""))
                ET.SubElement(el, "Service").text = str(line.service or "")
                ET.SubElement(el, "From").text = str(line.from_node or "")
                ET.SubElement(el, "To").text = str(line.to_node or "")
            
        # 3. Instrument Group
        if graph.instruments:
            inst_group = ET.SubElement(root, "Instrumentation")
            for inst in graph.instruments:
                el = ET.SubElement(inst_group, "Instrument", Tag=str(inst.tag or ""), Type=str(inst.type or ""))
                ET.SubElement(el, "LoopID").text = str(inst.loop_id or "")
            
        # 4. Valve Group
        if graph.valves:
            valve_group = ET.SubElement(root, "Valves")
            for valve in graph.valves:
                el = ET.SubElement(valve_group, "Valve", Tag=str(valve.tag or ""), Type=str(valve.type or ""), Size=str(valve.size or ""))
                ET.SubElement(el, "LineTag").text = str(valve.line_tag or "")
                ET.SubElement(el, "Rating").text = str(valve.rating or "")
                ET.SubElement(el, "NormalState").text = str(valve.normal_state or "")

        # 5. Safety Relief Valve Group
        if graph.safety_relief_valves:
            psv_group = ET.SubElement(root, "SafetyReliefValves")
            for psv in graph.safety_relief_valves:
                el = ET.SubElement(psv_group, "SafetyReliefValve", Tag=str(psv.tag or ""), Type=str(psv.type or ""), SetPressure=str(psv.set_pressure or ""))
                ET.SubElement(el, "Description").text = str(psv.service or "")
                ET.SubElement(el, "InletSize").text = str(psv.inlet_size or "")
                ET.SubElement(el, "OutletSize").text = str(psv.outlet_size or "")
                ET.SubElement(el, "InletSpec").text = str(psv.inlet_spec or "")
                ET.SubElement(el, "ReliefDestination").text = str(psv.relief_destination or "")

        # 6. Electrical Luminaires
        luminaires = getattr(graph, 'luminaires', [])
        if luminaires:
            lum_group = ET.SubElement(root, "Luminaires")
            for lum in luminaires:
                el = ET.SubElement(lum_group, "Luminaire", Tag=str(lum.tag or ""), FittingType=str(lum.fitting_type or ""))
                ET.SubElement(el, "Wattage").text = str(lum.wattage or "")
                ET.SubElement(el, "Circuit").text = str(lum.circuit or "")
                ET.SubElement(el, "Panel").text = str(lum.panel or "")

        # 7. Electrical Panels
        panels = getattr(graph, 'panels', [])
        if panels:
            panel_group = ET.SubElement(root, "ElectricalPanels")
            for pnl in panels:
                el = ET.SubElement(panel_group, "Panel", Tag=str(pnl.tag or ""), Type=str(pnl.panel_type or ""))
                ET.SubElement(el, "Voltage").text = str(pnl.voltage or "")
                ET.SubElement(el, "CapacityKVA").text = str(pnl.capacity_kva or "")
                ET.SubElement(el, "FeederFrom").text = str(pnl.feeder_from or "")

        # 8. Cables & Circuits
        cables = getattr(graph, 'cables', [])
        if cables:
            cable_group = ET.SubElement(root, "Cables")
            for cbl in cables:
                el = ET.SubElement(cable_group, "Cable", Tag=str(cbl.tag or ""), Type=str(cbl.cable_type or ""))
                ET.SubElement(el, "SizeMM2").text = str(cbl.size_mm2 or "")
                ET.SubElement(el, "FromPanel").text = str(cbl.from_panel or "")
                ET.SubElement(el, "ToEquipment").text = str(cbl.to_equipment or "")

        # 9. Earthing Components
        earthing = getattr(graph, 'earthing_components', [])
        if earthing:
            earth_group = ET.SubElement(root, "Earthing")
            for eth in earthing:
                el = ET.SubElement(earth_group, "EarthingComponent", Tag=str(eth.tag or ""), Type=str(eth.component_type or ""))
                ET.SubElement(el, "Material").text = str(eth.material or "")
                ET.SubElement(el, "ConnectedTo").text = str(eth.connected_to or "")

        # Prettify XML print output
        xml_str = ET.tostring(root, encoding="utf-8")
        parsed = minidom.parseString(xml_str)
        pretty_xml = parsed.toprettyxml(indent="  ")
        
        with open(path, "w", encoding="utf-8") as f:
            f.write(pretty_xml)

    def _generate_comos_json(self, graph, path: str):
        """
        Formats compiled graph data into a COMOS hierarchical plant object structure.
        COMOS structure: Project -> Unit -> Location -> Object.
        """
        comos_structure = {
            "comos_project": "SID_AI_EXTRACTS",
            "drawing_type": getattr(graph, 'drawing_type', 'GENERIC'),
            "units": [
                {
                    "unit_id": "UNIT_01",
                    "description": f"Extracted from {getattr(graph, 'drawing_type', 'ENGINEERING')} drawing",
                    "objects": []
                }
            ]
        }
        
        unit_objects = comos_structure["units"][0]["objects"]
        
        for eq in graph.equipment:
            unit_objects.append({
                "comos_class": "Equipment",
                "tag": eq.tag,
                "attributes": {
                    "Name": eq.name,
                    "Type": eq.type,
                    "Description": eq.description,
                    "Material": eq.material
                }
            })
            
        for line in graph.lines:
            unit_objects.append({
                "comos_class": "PipeLine",
                "tag": line.tag,
                "attributes": {
                    "NominalSize": line.size,
                    "ServiceCode": line.service,
                    "PipingSpec": line.spec,
                    "FromConnection": line.from_node,
                    "ToConnection": line.to_node
                }
            })
            
        for inst in graph.instruments:
            unit_objects.append({
                "comos_class": "InstrumentBubble",
                "tag": inst.tag,
                "attributes": {
                    "InstrumentType": inst.type,
                    "LoopID": inst.loop_id,
                    "Location": inst.location
                }
            })
            
        for valve in graph.valves:
            unit_objects.append({
                "comos_class": "ValveComponent",
                "tag": valve.tag,
                "attributes": {
                    "ValveType": valve.type,
                    "NominalSize": valve.size,
                    "NormalState": valve.normal_state,
                    "RatingClass": valve.rating,
                    "HostLine": valve.line_tag
                }
            })

        for psv in graph.safety_relief_valves:
            unit_objects.append({
                "comos_class": "SafetyReliefValve",
                "tag": psv.tag,
                "attributes": {
                    "ValveType": psv.type,
                    "Description": psv.service,
                    "SetPressure": psv.set_pressure,
                    "InletSize": psv.inlet_size,
                    "OutletSize": psv.outlet_size,
                    "InletSpec": psv.inlet_spec,
                    "ReliefDestination": psv.relief_destination
                }
            })

        # Electrical items
        for lum in getattr(graph, 'luminaires', []):
            unit_objects.append({"comos_class": "Luminaire", "tag": lum.tag,
                "attributes": {"FittingType": lum.fitting_type, "Wattage": lum.wattage,
                    "Circuit": lum.circuit, "Panel": lum.panel}})

        for panel in getattr(graph, 'panels', []):
            unit_objects.append({"comos_class": "ElectricalPanel", "tag": panel.tag,
                "attributes": {"PanelType": panel.panel_type, "Voltage": panel.voltage,
                    "Capacity": panel.capacity_kva, "FeederFrom": panel.feeder_from}})

        for cable in getattr(graph, 'cables', []):
            unit_objects.append({"comos_class": "Cable", "tag": cable.tag,
                "attributes": {"CableType": cable.cable_type, "SizeMM2": cable.size_mm2,
                    "Cores": cable.cores, "FromPanel": cable.from_panel, "ToEquipment": cable.to_equipment}})

        for earth in getattr(graph, 'earthing_components', []):
            unit_objects.append({"comos_class": "EarthingComponent", "tag": earth.tag,
                "attributes": {"ComponentType": earth.component_type, "Material": earth.material,
                    "ConnectedTo": earth.connected_to}})
            
        with open(path, "w") as f:
            json.dump(comos_structure, f, indent=2)

    def _generate_sppid_csv(self, graph, path: str):
        """
        Generates relational import CSV sheets suited for SmartPlant & Industrial DB import tables.
        Outputs exact extracted entities and real topological relationships (FEEDS, EARTHED_TO, MONITORS, INSTALLED_ON, CONNECTS_TO).
        """
        rows = []
        rows.append(["ENTITY_TYPE", "TAG", "PROPERTY_1", "PROPERTY_2", "PROPERTY_3", "PROPERTY_4", "PROPERTY_5"])

        # Entities
        for eq in graph.equipment:
            rows.append(["EQUIPMENT", eq.tag or "", eq.name or "", eq.type or "", eq.description or "", eq.location or "", ""])

        for line in graph.lines:
            rows.append(["PIPING_LINE", line.tag or "", line.size or "", line.service or "", line.spec or "", line.from_node or "", line.to_node or ""])

        for inst in graph.instruments:
            rows.append(["INSTRUMENT", inst.tag or "", inst.type or "", inst.loop_id or "", inst.location or "", "", ""])

        for valve in graph.valves:
            rows.append(["VALVE", valve.tag or "", valve.type or "", valve.size or "", valve.line_tag or "", valve.rating or "", valve.normal_state or ""])

        for psv in graph.safety_relief_valves:
            rows.append(["SAFETY_RELIEF_VALVE", psv.tag or "", psv.type or "", psv.set_pressure or "", psv.inlet_size or "", psv.outlet_size or "", psv.relief_destination or ""])

        for lum in getattr(graph, 'luminaires', []):
            rows.append(["LUMINAIRE", lum.tag or "", lum.fitting_type or "", lum.wattage or "", lum.circuit or "", lum.panel or "", lum.location or ""])

        for panel in getattr(graph, 'panels', []):
            rows.append(["ELECTRICAL_PANEL", panel.tag or "", panel.panel_type or "", panel.voltage or "", panel.capacity_kva or "", panel.feeder_from or "", panel.location or ""])

        for cable in getattr(graph, 'cables', []):
            rows.append(["CABLE", cable.tag or "", cable.cable_type or "", cable.size_mm2 or "", cable.cores or "", cable.from_panel or "", cable.to_equipment or ""])

        for earth in getattr(graph, 'earthing_components', []):
            rows.append(["EARTHING", earth.tag or "", earth.component_type or "", earth.material or "", earth.size or "", earth.connected_to or "", earth.resistance or ""])

        for gen in getattr(graph, 'generic_components', []):
            rows.append(["GENERIC_COMPONENT", gen.tag or "", gen.classification or "", gen.description or "", "", "", ""])

        # Topological Relationships (Fix Gemini CSV 0% score)
        for rel in getattr(graph, 'relationships', []):
            src_tag = getattr(rel, 'source_tag', getattr(rel, 'source', ''))
            trg_tag = getattr(rel, 'target_tag', getattr(rel, 'target', ''))
            r_type = str(getattr(rel, 'rel_type', getattr(rel, 'type', ''))).upper()
            conf_val = getattr(rel, 'confidence', 1.0)
            rows.append(["RELATIONSHIP", src_tag, r_type, trg_tag, f"conf={conf_val:.2f}", "", ""])

        df = pd.DataFrame(rows)
        df.to_csv(path, index=False, header=False)

